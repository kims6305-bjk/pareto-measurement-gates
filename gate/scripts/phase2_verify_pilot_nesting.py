"""검증 게이트: 파일럿이 정본 순서의 진짜 앞부분(nested)인지 실측."""
import json
import subprocess
from pathlib import Path

GATE = Path("/Users/bjkim/.openclaw/workspace/projects/probe-graph-public/gate/scripts")

canon = json.load(open(GATE / "phase2_canonical_order.json", encoding="utf-8"))
pilot = json.load(open(GATE / "phase2_pilot_label_sheet.json", encoding="utf-8"))

canon_ids = [r["id"] for r in canon]
pilot_ids = [r["id"] for r in pilot]

assert canon_ids[:len(pilot_ids)] == pilot_ids, "파일럿이 정본 순서 앞부분이 아님"
print(f"✅ 중첩성: 파일럿 {len(pilot_ids)}건 = 정본 {len(canon_ids)}건의 앞부분")

# 기존 본 시트(CAP=3)와 집합이 같은지 — 정본 순서가 본 표본 모집단과 동일해야 함
main = json.load(open(GATE / "phase2_human_label_sheet.json", encoding="utf-8"))
main_ids = {r["id"] for r in main}
if len(main_ids) == len(canon_ids):
    assert main_ids == set(canon_ids), "정본 순서와 본 시트 집합 불일치"
    print(f"✅ 모집단 일치: 본 시트 {len(main_ids)}건 == 정본 순서 {len(canon_ids)}건")
else:
    print(f"ℹ️ 본 시트는 CAP={len(main_ids)}건 설정으로 생성됨 "
          f"(정본 {len(canon_ids)}건) — 본 시트는 N 확정 후 재생성 예정")
    assert main_ids <= set(canon_ids), "본 시트가 정본 순서의 부분집합이 아님"
    print("✅ 부분집합 관계는 성립")

# 블라인드: 파일럿 xlsx에 gold/저지/프로브 열이 없는지
from openpyxl import load_workbook
ws = load_workbook(GATE / "phase2_pilot_label_sheet.xlsx")["phase2_pilot"]
head = [str(c.value) if c.value is not None else "" for c in ws[1]]
leak = [h for h in head if h and (h.startswith("_") or "gold" in h.lower()
                                  or "judge" in h.lower() or "probe" in h.lower())]
assert not leak, f"블라인드 위반 열: {leak}"
print(f"✅ 블라인드: 노출 열 {head}")
print(f"✅ 행 수: {ws.max_row - 1}건")

# 미라벨 상태에서 determine_n 이 거부하는지 (fail-closed)
r = subprocess.run(
    [".venv/bin/python", "scripts/phase2_determine_n.py"],
    cwd=GATE.parent, capture_output=True, text=True)
assert r.returncode != 0 and "미라벨" in (r.stdout + r.stderr), \
    f"미라벨 시트를 그냥 통과시킴: rc={r.returncode} {r.stdout} {r.stderr}"
print(f"✅ fail-closed: 미라벨 상태에서 N 산출 거부 — {(r.stdout+r.stderr).strip().splitlines()[-1]}")
