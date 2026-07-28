"""Phase 2 내부 파일럿(internal pilot) 시트 생성 — 기저율 추정 전용.

PHASE2_INTERNAL_PILOT.md 를 커밋한 **뒤에** 실행한다.

핵심 설계:
  정본 순서(canonical order)를 seed 고정으로 한 번 만들고, 파일럿 = 그 순서의 앞
  PILOT_N건, 본 표본 = 같은 순서의 앞 N건으로 정의한다. 따라서 파일럿은 본 표본에
  **중첩(nested)** 되며 폐기되지 않는다 — 이것이 파일럿 라벨을 버리지 않고도
  optional stopping 비판을 피하는 조건이다.

  파일럿은 SELECTION 규칙을 앞으로 당기지 **않는다.** R4(선택규칙 전량 포함)는
  본 표본 수준의 요구이고, 파일럿의 유일한 임무는 **편향 없는 기저율 추정**이다.
  선택규칙을 앞으로 당기면 그 층의 문제 발생률이 기저율 추정을 오염시킨다.
"""
import json
import os
import random
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

P = Path("<repo>")
GATE = P / "gate/scripts"
SEED = 20260728          # phase2_build_label_sheet.py 와 동일 시드
CAP = 3                  # 문항당 주장 캡 (SELECTION 면제) — 정본 순서의 모집단
PILOT_N = int(os.environ.get("PHASE2_PILOT_N", "30"))

cand = json.load(open(GATE / "phase2_candidate_questions.json", encoding="utf-8"))
results = json.load(open(P / "ab/ab_results.json", encoding="utf-8"))
frozen = {q["qid"]: q for q in
          json.load(open(P / "ab/ab_questions_FROZEN.json", encoding="utf-8"))["questions"]}

# --- 1. 정본 순서 만들기 (phase2_build_label_sheet.py 와 동일 절차) ---------
rows = []
for q in cand["questions"]:
    qid = q["qid"]
    res = results.get(qid)
    if not res or not isinstance(res.get("armA"), dict):
        continue
    claims = res["armA"].get("claims") or []
    if not claims:
        continue
    answer = res["armA"].get("answer", "")
    for c in claims:
        rows.append({
            "id": f"{qid}-c{c.get('claim_id')}",
            "qid": qid,
            "layer": q["layer"],
            "standard": q["standard"],
            "rule_type": q["rule_type"],
            "question": q["question"],
            "evidence": q["evidence_paragraphs"],
            "full_answer": answer,
            "claim_text": c.get("text", ""),
            "claim_citation": c.get("citation", ""),
            "_gold_answer": frozen[qid].get("gold_answer", ""),
            "_gold_citations": json.dumps(frozen[qid].get("gold_citations", []),
                                          ensure_ascii=False),
            "_trap_note": frozen[qid].get("trap_note", "") or "",
            "human_label": "",
            "human_note": "",
        })

random.Random(SEED).shuffle(rows)

kept, seen = [], Counter()
for r in rows:
    if r["rule_type"] == "SELECTION" or seen[r["qid"]] < CAP:
        kept.append(r)
        seen[r["qid"]] += 1
canonical = kept
for i, r in enumerate(canonical, start=1):
    r["order"] = i

(GATE / "phase2_canonical_order.json").write_text(
    json.dumps([{"order": r["order"], "id": r["id"], "qid": r["qid"],
                 "layer": r["layer"], "rule_type": r["rule_type"]}
                for r in canonical], ensure_ascii=False, indent=2),
    encoding="utf-8")

pilot = canonical[:PILOT_N]

print(f"정본 순서(CAP={CAP}, seed={SEED}): {len(canonical)}건 / 문항 "
      f"{len({r['qid'] for r in canonical})}개")
print(f"파일럿: 앞 {PILOT_N}건 = 문항 {len({r['qid'] for r in pilot})}개")
print("  층 분포:", dict(Counter(r["layer"] for r in pilot)))
print("  규칙유형:", dict(Counter(r["rule_type"] for r in pilot)))
print("  기준서 종수:", len({r["standard"] for r in pilot}))

# --- 2. 블라인드 xlsx ------------------------------------------------------
VISIBLE = [("id", 14), ("layer", 12), ("standard", 20), ("question", 46),
           ("evidence", 74), ("full_answer", 52), ("claim_text", 52),
           ("claim_citation", 20), ("human_label", 12), ("human_note", 34)]
wb = Workbook()
ws = wb.active
ws.title = "phase2_pilot"
head = PatternFill("solid", fgColor="141F38")
inputf = PatternFill("solid", fgColor="FDF2E3")
for i, (name, w) in enumerate(VISIBLE, start=1):
    cell = ws.cell(row=1, column=i, value=name)
    cell.font = Font(bold=True, color="FFFFFF", name="Apple SD Gothic Neo")
    cell.fill = head
    ws.column_dimensions[cell.column_letter].width = w
ws.freeze_panes = "A2"

for ri, r in enumerate(pilot, start=2):
    for ci, (name, _) in enumerate(VISIBLE, start=1):
        cell = ws.cell(row=ri, column=ci, value=r[name])
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.font = Font(name="Apple SD Gothic Neo", size=10)
        if name in ("human_label", "human_note"):
            cell.fill = inputf
    ws.row_dimensions[ri].height = 96

col_label = chr(ord("A") + VISIBLE.index(("human_label", 12)))
dv = DataValidation(type="list", formula1='"S,C,I"', allow_blank=True,
                    showDropDown=False)
dv.error = "S / C / I 중 하나만 입력"
ws.add_data_validation(dv)
dv.add(f"{col_label}2:{col_label}{len(pilot)+1}")

ws2 = wb.create_sheet("라벨링_기준")
guide = [
    ["Phase 2 파일럿 라벨링 — 기준 요약"],
    ["정본", "gate/PHASE2_LABELING_PROTOCOL.md + gate/PHASE2_INTERNAL_PILOT.md"],
    [""],
    ["S", "claim이 제공된 근거 문단에 명시돼 있거나 직접 도출됨"],
    ["C", "근거 문단이 claim과 반대되는 내용을 진술"],
    ["I", "근거만으로 확인 불가 — 회계적으로 옳아도 근거 범위 밖이면 I"],
    [""],
    ["핵심", "판단 기준은 '근거가 claim을 지지하는가'뿐. 도메인 지식으로 참/거짓 판정하지 않음"],
    [""],
    ["메모 규칙 (human_note)"],
    ["분해", "하나의 규칙이 여러 주장으로 쪼개진 경우 — 전체 답변에서 성립하면 S + 메모 '분해' (필수)"],
    ["선택분기", "근거가 '둘 중 이른 날' 류 선택 규칙인데 주장이 한 분기만 서술 — 메모 '선택분기' (필수)"],
    ["명제없음", "claim이 명제가 아님 — I + 메모 '명제없음'"],
    ["애매", "S/I 경계에서 애매하면 I로 기울이고 메모 한 줄"],
    [""],
    ["이 시트의 성격", "파일럿 30건은 버리는 표본이 아니라 본 표본의 앞 30건입니다. "
                    "여기서 읽는 것은 기저율(문제 사례 비율) 하나뿐이고, "
                    "프로브·저지 성능 지표는 열지 않습니다."],
    ["금지", "저지/프로브 출력 열람 금지. 라벨링 중 문항 수정 금지."],
]
for r_ in guide:
    ws2.append(r_)
ws2.column_dimensions["A"].width = 14
ws2.column_dimensions["B"].width = 96
for row in ws2.iter_rows():
    for c in row:
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.font = Font(name="Apple SD Gothic Neo", size=10)
ws2["A1"].font = Font(bold=True, size=12, name="Apple SD Gothic Neo")

out = GATE / "phase2_pilot_label_sheet.xlsx"
wb.save(out)
(GATE / "phase2_pilot_label_sheet.json").write_text(
    json.dumps(pilot, ensure_ascii=False, indent=2), encoding="utf-8")

print(f"\nsaved: {out.relative_to(P)}  ({out.stat().st_size/1024:.1f} KB)")
print(f"saved: {(GATE/'phase2_pilot_label_sheet.json').relative_to(P)}")
print(f"saved: {(GATE/'phase2_canonical_order.json').relative_to(P)} "
      f"(정본 순서 — N 확정의 근거)")
