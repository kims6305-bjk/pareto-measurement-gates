"""Phase 1: judge P/R scoring against human labels.

Run AFTER bjkim fills column G in phase1_human_label_sheet.xlsx.
Implements gate/LABELING_PROTOCOL.md:
  - human label = ground truth, target P/R >= 90%
  - '분해' memo rows -> dual-definition scoring (context vs isolated)
  - '명제없음' memo rows -> excluded from P/R, reported as extraction defects
"""
import json
import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

GATE = Path(__file__).resolve().parent.parent
SHEET = GATE / "scripts" / "phase1_human_label_sheet.xlsx"

# judge labels in sheet are gate semantic labels; map both scales to S/C/I
JUDGE_MAP = {"SUPPORTED": "S", "CONTRADICTED": "C", "INSUFFICIENT": "I",
             "S": "S", "C": "C", "I": "I"}
VALID_HUMAN = {"S", "C", "I"}


def prf(rows, positive: str):
    """Precision/recall for one class, judge vs human."""
    tp = sum(1 for h, j in rows if h == positive and j == positive)
    fp = sum(1 for h, j in rows if h != positive and j == positive)
    fn = sum(1 for h, j in rows if h == positive and j != positive)
    p = tp / (tp + fp) if tp + fp else None
    r = tp / (tp + fn) if tp + fn else None
    return {"tp": tp, "fp": fp, "fn": fn, "precision": p, "recall": r}


def main():
    wb = load_workbook(SHEET)
    ws = wb["라벨링"]
    records, unlabeled, invalid = [], [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rid, human, memo, judge = row[0], row[6], row[7] or "", row[8]
        if rid is None:
            continue
        human = (str(human).strip().upper() if human is not None else "")
        if not human:
            unlabeled.append(rid)
            continue
        if human not in VALID_HUMAN:
            invalid.append((rid, human))
            continue
        judge_n = JUDGE_MAP.get(str(judge).strip().upper())
        records.append({"id": rid, "human": human, "judge": judge_n,
                        "memo": str(memo).strip()})

    if unlabeled:
        print(f"⚠️ 미라벨 {len(unlabeled)}건: {unlabeled[:10]}{'...' if len(unlabeled) > 10 else ''}")
        print("전건 라벨 후 재실행. 중간 점검용 부분 집계는 아래에 계속 출력.")
    if invalid:
        print(f"❌ 유효하지 않은 라벨 {len(invalid)}건 (S/C/I만 허용): {invalid}")

    defects = [r for r in records if "명제없음" in r["memo"]]
    scored = [r for r in records if "명제없음" not in r["memo"]]
    decomposed = [r for r in scored if "분해" in r["memo"]]

    print(f"\n라벨 완료 {len(records)} / 채점 대상 {len(scored)} / "
          f"명제없음(제외) {len(defects)} / 분해 표시 {len(decomposed)}")
    print("사람 라벨 분포:", dict(Counter(r["human"] for r in scored)))
    print("저지 라벨 분포:", dict(Counter(r["judge"] for r in scored)))

    def report(name, pairs):
        agree = sum(1 for h, j in pairs if h == j)
        print(f"\n== {name} (n={len(pairs)}, agreement {agree}/{len(pairs)}"
              f" = {agree/len(pairs):.1%}) ==" if pairs else f"\n== {name} (n=0) ==")
        ok_all = True
        for cls in ("S", "C", "I"):
            m = prf(pairs, cls)
            fmt = lambda v: "n/a" if v is None else f"{v:.1%}"
            flag = ""
            for k in ("precision", "recall"):
                if m[k] is not None and m[k] < 0.90:
                    ok_all = False
                    flag = " ⚠️<90%"
            print(f"  {cls}: P={fmt(m['precision'])} R={fmt(m['recall'])} "
                  f"(tp{m['tp']}/fp{m['fp']}/fn{m['fn']}){flag}")
        # gate-level: judge positive = C or I (시비). 배포 관점 핵심 지표.
        pos_pairs = [("POS" if h in "CI" else "NEG", "POS" if j in "CI" else "NEG")
                     for h, j in pairs]
        m = prf(pos_pairs, "POS")
        fmt = lambda v: "n/a" if v is None else f"{v:.1%}"
        print(f"  [게이트 양성=C∪I] P={fmt(m['precision'])} R={fmt(m['recall'])}")
        return ok_all

    # 1차 정의: 맥락 기준 (분해 건 정답 = 시트의 사람 라벨 그대로)
    ctx_pairs = [(r["human"], r["judge"]) for r in scored]
    ctx_ok = report("맥락 기준 (1차 정의)", ctx_pairs)

    # 2차 정의: 고립 기준 — 분해 건은 claim 단독으론 근거 불충분 = I 로 재해석
    iso_pairs = [("I" if "분해" in r["memo"] else r["human"], r["judge"])
                 for r in scored]
    iso_ok = report("고립 기준 (분해→I 재라벨)", iso_pairs)

    print(f"\n판정: 맥락 기준 {'PASS' if ctx_ok else 'FAIL'} / "
          f"고립 기준 {'PASS' if iso_ok else 'FAIL'} (목표 P·R≥90%)")
    if ctx_ok != iso_ok:
        print("⚠️ 정의에 따라 결론이 뒤집힘 — LABELING_PROTOCOL.md에 따라 양쪽 병기 필수")
    if defects:
        print(f"\n추출 불량(명제없음): {[r['id'] for r in defects]}")

    out = GATE / "scripts" / "phase1_judge_pr_result.json"
    out.write_text(json.dumps({
        "n_labeled": len(records), "n_scored": len(scored),
        "unlabeled": unlabeled, "defects": [r["id"] for r in defects],
        "decomposed": [r["id"] for r in decomposed],
        "records": records}, ensure_ascii=False, indent=2))
    print("\nsaved:", out)
    return 0 if not unlabeled and not invalid else 1


if __name__ == "__main__":
    sys.exit(main())
