"""파일럿 라벨의 메모 태그를 집계한다 (기저율 외 부수 관찰 — 화이트리스트 §1.1-2 범위).

⚠️ 함정: 메모는 자유서술이라 단순 substring 매칭이 오탐을 낸다.
   실제 사례 — Q089-c3 "…다른 규칙이라 **분해 아님**. 인용 문단53에 명시되어 S"
   `'분해' in note` 로 세면 이 건이 분해로 잡힌다. 부정 표현을 먼저 제외해야 한다.
"""
import json
import re
from pathlib import Path

from openpyxl import load_workbook

GATE = Path("/Users/bjkim/.openclaw/workspace/projects/probe-graph-public/gate/scripts")
ws = load_workbook(GATE / "phase2_pilot_label_sheet.xlsx")["phase2_pilot"]

rows = [{"id": r[0], "label": str(r[8] or "").strip().upper(),
         "note": str(r[9] or "").strip()}
        for r in ws.iter_rows(min_row=2, values_only=True) if r[0]]

NEG = ("아님", "아니", "해당 없", "해당없", "제외")


def has_tag(note: str, tag: str) -> bool:
    """태그가 긍정으로 쓰였는지. '분해 아님' 류 부정 문맥은 제외한다."""
    for m in re.finditer(re.escape(tag), note):
        tail = note[m.end():m.end() + 12]
        if any(neg in tail for neg in NEG):
            continue
        return True
    return False


tags = {}
for tag in ("분해", "선택분기", "명제없음"):
    hit = [r["id"] for r in rows if has_tag(r["note"], tag)]
    naive = [r["id"] for r in rows if tag in r["note"]]
    tags[tag] = hit
    fp = sorted(set(naive) - set(hit))
    print(f"{tag}: {len(hit)}건 {hit}")
    if fp:
        print(f"   (단순매칭 오탐 제외됨: {fp})")

problems = [r for r in rows if r["label"] in ("C", "I")
            and not has_tag(r["note"], "명제없음")]
print(f"\n문제 사례(C/I): {len(problems)}건 "
      f"{[(r['id'], r['label']) for r in problems]}")
print(f"메모 기재율: {sum(1 for r in rows if r['note'])}/{len(rows)}")

out = GATE / "phase2_pilot_tags.json"
out.write_text(json.dumps({
    "n": len(rows),
    "tags": {k: v for k, v in tags.items()},
    "problems": [{"id": r["id"], "label": r["label"], "note": r["note"]}
                 for r in problems],
    "note": "부정문맥('분해 아님') 제외 후 집계. 단순 substring 매칭은 오탐을 낸다.",
}, ensure_ascii=False, indent=2), encoding="utf-8")
print(f"\nsaved: {out.name}")
