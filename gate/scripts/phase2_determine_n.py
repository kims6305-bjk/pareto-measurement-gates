"""파일럿 라벨에서 기저율만 읽어 본 표본 N을 기계적으로 확정한다.

PHASE2_INTERNAL_PILOT.md §3 규칙을 코드로 옮긴 것이며, 재량 파라미터가 없다.
효과 지표(프로브/저지 출력)는 이 스크립트가 **읽지도 않는다** — 화이트리스트를
코드 수준에서 강제하기 위해 해당 파일 경로를 아예 참조하지 않는다.

사용:
    .venv/bin/python scripts/phase2_determine_n.py
        (기본: phase2_pilot_label_sheet.xlsx 의 human_label / human_note 열)
"""
import json
import math
from pathlib import Path

from openpyxl import load_workbook

P = Path("/Users/bjkim/.openclaw/workspace/projects/probe-graph-public")
GATE = P / "gate/scripts"

TARGET_PROBLEMS = 55   # phase2_power_analysis.py 실측
N_MIN, N_MAX = 90, None  # N_MAX = 정본 순서 전량 (아래에서 결정)

sheet = GATE / "phase2_pilot_label_sheet.xlsx"
wb = load_workbook(sheet)
ws = wb["phase2_pilot"]
head = [c.value for c in ws[1]]
i_id, i_lab, i_note = (head.index("id"), head.index("human_label"),
                       head.index("human_note"))

labels = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[i_id] is None:
        continue
    labels.append({
        "id": row[i_id],
        "label": str(row[i_lab] or "").strip().upper(),
        "note": str(row[i_note] or "").strip(),
    })

blank = [r["id"] for r in labels if not r["label"]]
if blank:
    raise SystemExit(f"미라벨 {len(blank)}건 남음 — 전량 라벨 후 실행: {blank[:5]}")

bad = [r["id"] for r in labels if r["label"] not in ("S", "C", "I")]
if bad:
    raise SystemExit(f"S/C/I 외 라벨 {len(bad)}건: {bad[:5]}")

no_prop = [r for r in labels if "명제없음" in r["note"]]
denom = len(labels) - len(no_prop)
problems = [r for r in labels
            if r["label"] in ("C", "I") and "명제없음" not in r["note"]]

if denom <= 0:
    raise SystemExit("분모 0 — 파일럿 전건이 명제없음. 문항 재설계 필요.")

p_hat = len(problems) / denom

canonical = json.load(open(GATE / "phase2_canonical_order.json", encoding="utf-8"))
N_MAX = len(canonical)

if p_hat == 0:
    n_req = N_MAX
else:
    n_req = math.ceil(TARGET_PROBLEMS / p_hat)
N = max(N_MIN, min(n_req, N_MAX))
shortfall = n_req > N_MAX

# 군집 계수 — 열람 화이트리스트 §1.1-2
qids = {r["id"].rsplit("-c", 1)[0] for r in problems}

print("=== 파일럿 기저율 (열람 화이트리스트 내 항목만) ===")
print(f"라벨 총수 {len(labels)}건 / 명제없음 {len(no_prop)}건 -> 분모 {denom}건")
print(f"문제 사례(C 또는 I) {len(problems)}건 -> 기저율 p̂ = {p_hat:.3f}")
print(f"문제가 걸친 서로 다른 답변 {len(qids)}개")
print()
print("=== N 확정 (PHASE2_INTERNAL_PILOT.md §3, 재량 없음) ===")
print(f"N_req = ceil({TARGET_PROBLEMS} / {p_hat:.3f}) = {n_req}")
print(f"clamp({n_req}, {N_MIN}, {N_MAX}) -> **N = {N}**")
if shortfall:
    print(f"⚠️ N_req {n_req} > 후보 풀 전량 {N_MAX} — 목표 문제 사례 {TARGET_PROBLEMS}건 "
          f"미달 가능. 결과에 '판정 불가 가능'을 명기한다. 문항 증설은 별도 사전 선언.")
print()
print(f"다음: PHASE2_CLAIM_CAP=3 PHASE2_N={N} 으로 본 시트 생성. "
      f"파일럿 {len(labels)}건은 정본 순서 앞부분이므로 이월(재라벨 금지).")

out = GATE / "phase2_pilot_baserate.json"
out.write_text(json.dumps({
    "source": sheet.name,
    "n_labeled": len(labels),
    "n_no_proposition": len(no_prop),
    "denominator": denom,
    "n_problems": len(problems),
    "base_rate": round(p_hat, 4),
    "distinct_answers_with_problem": len(qids),
    "target_problems": TARGET_PROBLEMS,
    "n_required_raw": n_req,
    "n_min": N_MIN,
    "n_max_pool": N_MAX,
    "N_final": N,
    "shortfall_possible": shortfall,
    "note": "효과 지표(프로브/저지 출력)는 이 산출에 일절 사용되지 않음.",
}, ensure_ascii=False, indent=2), encoding="utf-8")
print(f"\nsaved: {out.relative_to(P)}")
