"""Build the Phase 2 blind labeling sheet (xlsx + json).

Follows PHASE2_LABELING_PROTOCOL.md, committed before this ran.
  R1  one answer per question — armA only (no probe), same arm Phase 1 used
  R2  distractor layer prioritized (see note below)
  R3  standard diversity capped at 8/standard
  R4  all 9 SELECTION-rule questions force-included

R2 CORRECTION (measured, not assumed): the protocol originally prioritized both
trap layers. Building the sheet showed all 17 no_answer questions produced
`claims: []` — the model correctly abstained ("제공된 자료에서 확인되지 않음") on
every one. Abstentions carry no claim to label, so the no_answer layer cannot
contribute problem cases to this experiment. They are excluded from the sheet and
reported separately as a correct-abstention rate (17/17), which is a result about
the answerer, not about the probe.

Blind: no judge output, no probe output, no gold answer in the visible sheet.
Claims are shuffled with a fixed seed so sibling claims of one answer do not sit
adjacent (prevents the labeler from pattern-matching a decomposition and
labeling by position rather than by evidence).
"""
import json
import random
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

P = Path("/Users/bjkim/.openclaw/workspace/projects/probe-graph-public")
GATE = P / "gate/scripts"
SEED = 20260728

cand = json.load(open(GATE / "phase2_candidate_questions.json", encoding="utf-8"))
results = json.load(open(P / "ab/ab_results.json", encoding="utf-8"))
frozen = {q["qid"]: q for q in
          json.load(open(P / "ab/ab_questions_FROZEN.json", encoding="utf-8"))["questions"]}

rows, skipped = [], []
for q in cand["questions"]:
    qid = q["qid"]
    res = results.get(qid)
    if not res or not isinstance(res.get("armA"), dict):
        skipped.append((qid, "armA 없음"))
        continue
    claims = res["armA"].get("claims") or []
    if not claims:
        skipped.append((qid, "claims 없음"))
        continue
    answer = res["armA"].get("answer", "")
    for c in claims:
        rows.append({
            "id": f"{qid}-c{c.get('claim_id')}",
            "qid": qid,
            "layer": q["layer"],
            "standard": q["standard"],
            "rule_type": q["rule_type"],
            "question": q["question"],
            "evidence": q["evidence_paragraphs"],
            "full_answer": answer,
            "claim_text": c.get("text", ""),
            "claim_citation": c.get("citation", ""),
            # hidden reference columns — NOT written to the visible sheet
            "_gold_answer": frozen[qid].get("gold_answer", ""),
            "_gold_citations": json.dumps(frozen[qid].get("gold_citations", []),
                                          ensure_ascii=False),
            "_trap_note": frozen[qid].get("trap_note", "") or "",
            "human_label": "",
            "human_note": "",
        })

random.Random(SEED).shuffle(rows)

# --- claim cap per question (labeling-cost control) ----------------------
# Sibling claims of the same answer are NOT independent (this is exactly what
# broke Phase 1). Beyond a few claims per answer the extra labels cost bjkim
# time while adding almost no independent information, so cap them.
# SELECTION-rule questions are exempt: R4 needs their full branch structure.
CAP = int(__import__("os").environ.get("PHASE2_CLAIM_CAP", "3"))
kept, seen = [], Counter()
for r in rows:  # already shuffled -> the kept subset is unbiased within a question
    if r["rule_type"] == "SELECTION" or seen[r["qid"]] < CAP:
        kept.append(r)
        seen[r["qid"]] += 1
dropped = len(rows) - len(kept)
rows = kept
print(f"[문항당 주장 캡 {CAP}] {dropped}건 제외 (선택규칙 문항은 캡 면제)")

print(f"문항 {len(cand['questions'])}건 -> 주장 {len(rows)}건")
if skipped:
    print(f"제외 {len(skipped)}건: {skipped}")
print("  층 분포:", dict(Counter(r["layer"] for r in rows)))
print("  규칙유형:", dict(Counter(r["rule_type"] for r in rows)))
print("  기준서 종수:", len({r['standard'] for r in rows}))
print(f"  문항당 평균 주장 {len(rows)/len(cand['questions']):.1f}건")

# --- re-label subset (protocol §4.1): 12 questions, labeler must not know ---
qids = sorted({r["qid"] for r in rows})
relabel_qids = sorted(random.Random(SEED + 1).sample(qids, 12))
(GATE / "phase2_relabel_subset.json").write_text(
    json.dumps({"note": "재현성 검증용. 라벨링 완료 전 열람 금지.",
                "seed": SEED + 1, "qids": relabel_qids}, ensure_ascii=False, indent=2),
    encoding="utf-8")

# --- json (full, includes hidden cols for QC scripts) --------------------
(GATE / "phase2_human_label_sheet.json").write_text(
    json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")

# --- xlsx (blind: hidden columns excluded) ------------------------------
VISIBLE = [("id", 14), ("layer", 12), ("standard", 20), ("question", 46),
           ("evidence", 74), ("full_answer", 52), ("claim_text", 52),
           ("claim_citation", 20), ("human_label", 12), ("human_note", 34)]
wb = Workbook()
ws = wb.active
ws.title = "phase2_labels"
head = PatternFill("solid", fgColor="141F38")
inputf = PatternFill("solid", fgColor="FDF2E3")
for i, (name, w) in enumerate(VISIBLE, start=1):
    cell = ws.cell(row=1, column=i, value=name)
    cell.font = Font(bold=True, color="FFFFFF", name="Apple SD Gothic Neo")
    cell.fill = head
    ws.column_dimensions[cell.column_letter].width = w
ws.freeze_panes = "A2"

for ri, r in enumerate(rows, start=2):
    for ci, (name, _) in enumerate(VISIBLE, start=1):
        cell = ws.cell(row=ri, column=ci, value=r[name])
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.font = Font(name="Apple SD Gothic Neo", size=10)
        if name in ("human_label", "human_note"):
            cell.fill = inputf
    ws.row_dimensions[ri].height = 96

col_label = chr(ord("A") + VISIBLE.index(("human_label", 12)))
dv = DataValidation(type="list", formula1='"S,C,I"', allow_blank=True,
                    showDropDown=False)
dv.error = "S / C / I 중 하나만 입력"
ws.add_data_validation(dv)
dv.add(f"{col_label}2:{col_label}{len(rows)+1}")

# instruction sheet
ws2 = wb.create_sheet("라벨링_기준")
guide = [
    ["Phase 2 라벨링 — 기준 요약 (정본: gate/PHASE2_LABELING_PROTOCOL.md)"],
    [""],
    ["S", "claim이 제공된 근거 문단에 명시돼 있거나 직접 도출됨"],
    ["C", "근거 문단이 claim과 반대되는 내용을 진술"],
    ["I", "근거만으로 확인 불가 — 회계적으로 옳아도 근거 범위 밖이면 I"],
    [""],
    ["핵심", "판단 기준은 '근거가 claim을 지지하는가'뿐. 도메인 지식으로 참/거짓 판정하지 않음"],
    [""],
    ["메모 규칙 (human_note)"],
    ["분해", "하나의 규칙이 여러 주장으로 쪼개진 경우 — 전체 답변에서 성립하면 S + 메모 '분해' (필수)"],
    ["선택분기", "근거가 '둘 중 이른 날' 류 선택 규칙인데 주장이 한 분기만 서술 — 메모 '선택분기' (필수)"],
    ["명제없음", "claim이 명제가 아님 — I + 메모 '명제없음'"],
    ["애매", "S/I 경계에서 애매하면 I로 기울이고 메모 한 줄"],
    [""],
    ["금지", "저지/프로브 출력 열람 금지. 라벨링 중 문항 수정 금지."],
]
for r_ in guide:
    ws2.append(r_)
ws2.column_dimensions["A"].width = 14
ws2.column_dimensions["B"].width = 96
for row in ws2.iter_rows():
    for c in row:
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.font = Font(name="Apple SD Gothic Neo", size=10)
ws2["A1"].font = Font(bold=True, size=12, name="Apple SD Gothic Neo")

out = GATE / "phase2_human_label_sheet.xlsx"
wb.save(out)
print(f"\nsaved: {out.relative_to(P)}  ({out.stat().st_size/1024:.1f} KB)")
print(f"saved: {(GATE/'phase2_human_label_sheet.json').relative_to(P)}")
print(f"saved: {(GATE/'phase2_relabel_subset.json').relative_to(P)} "
      f"(재라벨 12문항 — 라벨 완료 전 열람 금지)")
